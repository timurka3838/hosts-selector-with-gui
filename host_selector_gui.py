#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
table_tool.py  —  Обработка таблиц (v3)
=========================================
Этап 1 — «Обработка файлов»
  • Выбрать папку → удалить столбцы → собрать хосты → hosts_result.xlsx

Этап 2 — «Управление хостами»
  • Дерево: имя файла (свёрнуто) → список хостов
  • Сценарий А  — ввести замену для каждого файла вручную (с валидацией)
  • Сценарий Б  — добавить колонку «dang»
  • Удалить     — удалить строки с выбранными хостами

Зависимости:  pip install pandas openpyxl xlrd
"""

import os, re, time, logging, threading
from pathlib import Path
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Callable, Dict, List, Optional, Set, Tuple
from tkinter import filedialog, scrolledtext

import tkinter as tk
from tkinter import ttk, messagebox
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════════════
#  НАСТРОЙКИ СТОЛБЦОВ — редактируйте под свои данные
# ══════════════════════════════════════════════════════════════════════════

# Столбцы для УДАЛЕНИЯ — все совпадения убираются из файлов.
# Добавляйте любые дополнительные имена.
DROP_COLUMNS: List[str] = [
    'n3', 'n4', 'n7', 'n8',
]

# Возможные названия столбца с ХОСТАМИ — используется первое найденное.
# Добавляйте варианты написания, которые встречаются в ваших файлах.
HOST_COLUMN_NAMES: List[str] = [
    'host', 'hostname', 'Host', 'HOST', 'Hostname', 'HostName',
]

# ══════════════════════════════════════════════════════════════════════════
#  ПРОЧИЕ НАСТРОЙКИ
# ══════════════════════════════════════════════════════════════════════════
CHUNK_SIZE      = 100_000          # строк за один чанк при обработке CSV
MAX_WORKERS     = os.cpu_count() or 4   # потоков (каждый файл — отдельный поток)
OUTPUT_FILENAME = 'hosts_result.xlsx'
EMPTY_DISPLAY   = '(пусто)'            # метка для пустого значения хоста

# ══════════════════════════════════════════════════════════════════════════
#  Логирование
# ══════════════════════════════════════════════════════════════════════════
log = logging.getLogger('table_tool')
log.setLevel(logging.DEBUG)
_fh = logging.FileHandler('table_tool.log', encoding='utf-8')
_fh.setFormatter(logging.Formatter('%(asctime)s  %(levelname)-8s  %(message)s'))
log.addHandler(_fh)
_sh = logging.StreamHandler()
_sh.setFormatter(logging.Formatter('%(asctime)s  %(levelname)-8s  %(message)s'))
log.addHandler(_sh)


# ══════════════════════════════════════════════════════════════════════════
#  Утилиты
# ══════════════════════════════════════════════════════════════════════════

def find_host_col_df(df) -> Optional[str]:
    """Имя столбца с хостами в DataFrame (первое из HOST_COLUMN_NAMES)."""
    for name in HOST_COLUMN_NAMES:
        if name in df.columns:
            return name
    return None


def find_host_idx(hdr: List[str]) -> Optional[int]:
    """Индекс столбца хостов в списке заголовков."""
    for name in HOST_COLUMN_NAMES:
        if name in hdr:
            return hdr.index(name)
    return None


def get_drop_cols_df(df) -> List[str]:
    """Список существующих столбцов для удаления из DataFrame."""
    drop_set = set(DROP_COLUMNS)
    return [c for c in df.columns if str(c) in drop_set]


def get_drop_idx(hdr: List[str]) -> Set[int]:
    """Индексы столбцов для удаления в списке заголовков."""
    drop_set = set(DROP_COLUMNS)
    return {i for i, h in enumerate(hdr) if h in drop_set}


def detect_csv_encoding(path: Path) -> str:
    for enc in ('utf-8-sig', 'utf-8', 'cp1251', 'latin-1'):
        try:
            pd.read_csv(path, nrows=2, encoding=enc)
            return enc
        except Exception:
            continue
    return 'utf-8'


def _remove(path: Path) -> None:
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


def get_source_files(folder_path: str) -> List[Path]:
    """Все .csv/.xlsx/.xlsm в папке, кроме hosts_result.xlsx."""
    return sorted(
        p for p in Path(folder_path).iterdir()
        if p.is_file()
        and p.suffix.lower() in {'.csv', '.xlsx', '.xlsm'}
        and p.name != OUTPUT_FILENAME
    )


def is_file_locked(path: Path) -> bool:
    """Возвращает True, если файл открыт другой программой."""
    # Office создаёт скрытый ~$filename при открытии
    lock = path.parent / f'~${path.name}'
    if lock.exists():
        return True
    # Пробуем открыть в эксклюзивном режиме
    try:
        with open(path, 'r+b'):
            pass
        return False
    except (IOError, OSError, PermissionError):
        return True


def check_file_locks(folder_path: str) -> List[str]:
    """Возвращает список имён файлов, которые заблокированы."""
    locked = []
    for p in Path(folder_path).iterdir():
        if (p.is_file()
                and p.suffix.lower() in {'.csv', '.xls', '.xlsx', '.xlsm'}
                and p.name != OUTPUT_FILENAME
                and is_file_locked(p)):
            locked.append(p.name)
    return sorted(locked)


# ══════════════════════════════════════════════════════════════════════════
#  Этап 1: обработка исходных файлов
#
#  Производительность:
#   • CSV   — pd.read_csv(chunksize=CHUNK_SIZE): читает и пишет CHUNK_SIZE строк
#             за раз, файл никогда полностью не в памяти.
#   • XLSX  — openpyxl read_only=True + write_only=True: потоковое чтение/запись,
#             строка за строкой, без буферизации всего файла.
#   • XLS   — формат ограничен ≈65 тыс. строк, загружается целиком.
#   • Параллелизм — каждый файл обрабатывается в отдельном потоке
#             (ThreadPoolExecutor, MAX_WORKERS потоков).
# ══════════════════════════════════════════════════════════════════════════

def process_csv(filepath: str, fname: str) -> List[Tuple[str, str]]:
    """
    CSV читается итератором чанков по CHUNK_SIZE строк.
    В лог выводится номер каждого чанка и диапазон строк.
    """
    path  = Path(filepath)
    tmp   = path.with_suffix('.tmp.csv')
    hosts: Set[str] = set()
    enc   = detect_csv_encoding(path)
    t0    = time.perf_counter()
    chunk_num = 0
    rows_read = 0

    try:
        first = True
        for chunk in pd.read_csv(path, chunksize=CHUNK_SIZE,
                                  encoding=enc, low_memory=False):
            chunk_num += 1
            r0 = rows_read + 1
            rows_read += len(chunk)
            log.debug(f"  ↳ {fname}: чанк {chunk_num}  "
                      f"(строки {r0:,} – {rows_read:,})")

            hcol = find_host_col_df(chunk)
            if hcol:
                hosts.update(chunk[hcol].fillna('').astype(str).unique())

            drop = get_drop_cols_df(chunk)
            if drop:
                chunk.drop(columns=drop, inplace=True)

            chunk.to_csv(tmp, index=False,
                         mode='w' if first else 'a',
                         header=first, encoding='utf-8-sig')
            first = False

        os.replace(tmp, path)
        t = time.perf_counter() - t0
        log.info(f"[CSV]   {fname}  |  {chunk_num} чанк(ов)  |  "
                 f"{rows_read:,} строк  |  {len(hosts)} хостов  |  {t:.2f}с")
    except PermissionError:
        _remove(tmp)
        raise PermissionError(
            f"Файл занят другой программой (закройте и повторите): {fname}")
    except Exception:
        _remove(tmp)
        raise

    return [(fname, h) for h in hosts]


def process_xlsx(filepath: str, fname: str) -> List[Tuple[str, str]]:
    """
    XLSX: потоковое чтение/запись через openpyxl (O(1) памяти).
    Строки дополняются до ширины заголовка (sparse-формат xlsx может
    опускать хвостовые пустые ячейки, что вызывает IndexError).
    Прогресс выводится каждые CHUNK_SIZE строк.
    """
    path  = Path(filepath)
    tmp   = path.with_suffix('.tmp.xlsx')
    hosts: Set[str] = set()
    t0    = time.perf_counter()
    rows_read = 0

    try:
        wb_r = load_workbook(path, read_only=True, data_only=True)
        ws_r = wb_r.active
        it   = ws_r.iter_rows(values_only=True)

        hdr_raw  = list(next(it))
        hdr_len  = len(hdr_raw)
        hdr      = [str(h) if h is not None else '' for h in hdr_raw]
        drop_i   = get_drop_idx(hdr)
        host_i   = find_host_idx(hdr)
        new_hdr  = [h for i, h in enumerate(hdr_raw) if i not in drop_i]

        wb_w = openpyxl.Workbook(write_only=True)
        ws_w = wb_w.create_sheet()
        ws_w.append(new_hdr)

        for row in it:
            rows_read += 1
            # Дополняем разреженные строки до длины заголовка
            r = list(row)
            if len(r) < hdr_len:
                r.extend([None] * (hdr_len - len(r)))

            if host_i is not None:
                h = r[host_i]
                hosts.add('' if h is None else str(h))

            ws_w.append([v for i, v in enumerate(r) if i not in drop_i])

            if rows_read % CHUNK_SIZE == 0:
                log.debug(f"  ↳ {fname}: {rows_read:,} строк  "
                          f"({time.perf_counter()-t0:.1f}с)")

        wb_r.close()
        wb_w.save(tmp)
        os.replace(tmp, path)
        t = time.perf_counter() - t0
        log.info(f"[XLSX]  {fname}  |  {rows_read:,} строк  |  "
                 f"{len(hosts)} хостов  |  {t:.2f}с")
    except PermissionError:
        _remove(tmp)
        raise PermissionError(
            f"Файл занят другой программой (закройте и повторите): {fname}")
    except Exception:
        _remove(tmp)
        raise

    return [(fname, h) for h in hosts]


def process_xls(filepath: str, fname: str) -> List[Tuple[str, str]]:
    """XLS: формат ограничен ≈65 тыс. строк, поэтому загружаем целиком."""
    path  = Path(filepath)
    hosts: Set[str] = set()
    t0    = time.perf_counter()

    try:
        df = pd.read_excel(path, engine='xlrd')
        hcol = find_host_col_df(df)
        if hcol:
            hosts.update(df[hcol].fillna('').astype(str).unique())

        drop = get_drop_cols_df(df)
        if drop:
            df.drop(columns=drop, inplace=True)

        new_path = path.with_suffix('.xlsx')
        df.to_excel(new_path, index=False)
        path.unlink()
        t = time.perf_counter() - t0
        log.info(f"[XLS]   {path.name} → {new_path.name}  |  "
                 f"{len(df):,} строк  |  {len(hosts)} хостов  |  {t:.2f}с")
    except Exception:
        raise

    return [(fname, h) for h in hosts]


def process_file(filepath: str) -> List[Tuple[str, str]]:
    p    = Path(filepath)
    ext  = p.suffix.lower()
    fname = p.name      # полное имя файла (включая расширение)
    log.info(f"Начало обработки: {fname}")

    if ext == '.csv':               return process_csv(filepath, fname)
    elif ext in ('.xlsx', '.xlsm'): return process_xlsx(filepath, fname)
    elif ext == '.xls':             return process_xls(filepath, fname)
    else:
        log.warning(f"Пропущен (неизвестный формат): {fname}")
        return []


def run_phase1(folder_path: str,
               progress_cb: Optional[Callable] = None) -> List[Tuple[str, str]]:
    """
    Запускает этап 1 параллельно (ThreadPoolExecutor).
    Каждый файл — отдельный поток, обрабатывается независимо.
    """
    files = sorted(
        p for p in Path(folder_path).iterdir()
        if p.is_file()
        and p.suffix.lower() in {'.csv', '.xls', '.xlsx', '.xlsm'}
        and p.name != OUTPUT_FILENAME
    )
    if not files:
        raise FileNotFoundError(f'Файлы не найдены в: {folder_path}')

    log.info(f"Найдено файлов для обработки: {len(files)}  "
             f"(потоков: {min(MAX_WORKERS, len(files))})")

    all_rows: List[Tuple[str, str]] = []
    done = 0
    t0   = time.perf_counter()

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futs = {pool.submit(process_file, str(f)): f for f in files}
        for fut in as_completed(futs):
            fname = futs[fut].name
            done += 1
            try:
                rows = fut.result()
                all_rows.extend(rows)
            except Exception as exc:
                log.error(f"Ошибка [{done}/{len(files)}] {fname}: {exc}",
                           exc_info=True)
            if progress_cb:
                progress_cb(done, len(files), fname)

    log.info(f"Этап 1 завершён: {len(files)} файлов, "
             f"{len(all_rows)} хостов, {time.perf_counter()-t0:.2f}с")
    return all_rows


def save_hosts_xlsx(all_rows: List[Tuple[str, str]], out: Path) -> None:
    df = (
        pd.DataFrame(all_rows, columns=['Название файла', 'Имя хоста'])
        .drop_duplicates()
        .sort_values(['Название файла', 'Имя хоста'],
                     key=lambda s: s.str.lower(), ignore_index=True)
    )
    df.insert(0, 'Запись',
              df['Название файла'] + ' - ' +
              df['Имя хоста'].apply(lambda h: EMPTY_DISPLAY if h == '' else h))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Хосты'
    hfill = PatternFill('solid', start_color='1F4E79')
    hfont = Font(bold=True, color='FFFFFF', name='Arial', size=11)
    cfont = Font(name='Arial', size=10)
    left  = Alignment(horizontal='left',   vertical='center')
    ctr   = Alignment(horizontal='center', vertical='center')

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(1, ci, col)
        c.fill = hfill; c.font = hfont; c.alignment = ctr

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(ri, ci, val)
            c.font = cfont; c.alignment = left

    for ci, col in enumerate(df.columns, 1):
        w = max(len(str(col)), df[col].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(w + 4, 80)

    ws.freeze_panes = 'A2'
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    log.info(f"hosts_result.xlsx сохранён: {out}  ({len(df)} строк)")


# ══════════════════════════════════════════════════════════════════════════
#  Вспомогательные функции модификации файлов
# ══════════════════════════════════════════════════════════════════════════

def _rewrite_csv(fp: Path, transform) -> None:
    """Перезапись CSV чанками. transform(chunk) -> chunk | None."""
    enc = detect_csv_encoding(fp)
    tmp = fp.with_suffix('.tmp.csv')
    try:
        first     = True
        wrote_any = False
        chunk_num = 0
        rows_done = 0
        t0 = time.perf_counter()
        for chunk in pd.read_csv(fp, chunksize=CHUNK_SIZE,
                                  encoding=enc, low_memory=False):
            chunk_num += 1
            r0 = rows_done + 1
            rows_done += len(chunk)
            log.debug(f"  ↳ {fp.name}: чанк {chunk_num}  "
                      f"(строки {r0:,} – {rows_done:,})")
            chunk = transform(chunk)
            if chunk is None or chunk.empty:
                continue
            chunk.to_csv(tmp, index=False,
                         mode='w' if first else 'a',
                         header=first, encoding='utf-8-sig')
            first = False
            wrote_any = True
        if wrote_any:
            os.replace(tmp, fp)
        else:
            pd.read_csv(fp, nrows=0, encoding=enc).to_csv(
                fp, index=False, encoding='utf-8-sig')
            _remove(tmp)
    except PermissionError:
        _remove(tmp)
        raise PermissionError(
            f"Файл занят другой программой (закройте и повторите): {fp.name}")
    except Exception:
        _remove(tmp); raise


def _rewrite_xlsx(fp: Path, transform_row, new_headers_fn=None) -> None:
    """
    Потоковая перезапись XLSX.
    Строки дополняются до длины заголовка — исправляет IndexError на sparse-файлах
    (xlsx может не хранить хвостовые пустые ячейки, тогда row короче hdr).
    """
    tmp = fp.with_suffix('.tmp.xlsx')
    try:
        wb_r = load_workbook(fp, read_only=True, data_only=True)
        ws_r = wb_r.active
        it   = ws_r.iter_rows(values_only=True)
        hdr_raw = list(next(it))
        hdr_len = len(hdr_raw)
        hdr     = [str(h) if h is not None else '' for h in hdr_raw]
        new_hdr = new_headers_fn(hdr) if new_headers_fn else list(hdr_raw)

        wb_w = openpyxl.Workbook(write_only=True)
        ws_w = wb_w.create_sheet()
        ws_w.append(new_hdr)

        rows_done = 0
        t0 = time.perf_counter()
        for row in it:
            r = list(row)
            # Padding: sparse xlsx omits trailing empty cells
            if len(r) < hdr_len:
                r.extend([None] * (hdr_len - len(r)))
            r = transform_row(r, hdr)
            if r is not None:
                ws_w.append(r)
            rows_done += 1
            if rows_done % CHUNK_SIZE == 0:
                log.debug(f"  ↳ {fp.name}: {rows_done:,} строк  "
                          f"({time.perf_counter()-t0:.1f}с)")

        wb_r.close()
        wb_w.save(tmp)
        os.replace(tmp, fp)
    except PermissionError:
        _remove(tmp)
        raise PermissionError(
            f"Файл занят другой программой (закройте и повторите): {fp.name}")
    except Exception:
        _remove(tmp); raise


# ══════════════════════════════════════════════════════════════════════════
#  Сценарий А: заменить все хосты на заданное значение
# ══════════════════════════════════════════════════════════════════════════

def _replace_csv(fp: Path, val: str) -> None:
    def tr(chunk):
        hcol = find_host_col_df(chunk)
        if hcol:
            chunk[hcol] = val
        return chunk
    _rewrite_csv(fp, tr)


def _replace_xlsx(fp: Path, val: str) -> None:
    def tr(row, hdr):
        hi = find_host_idx(hdr)
        if hi is not None:
            row[hi] = val
        return row
    _rewrite_xlsx(fp, tr)


def scenario_a_apply(folder_path: str,
                     replacements: Dict[str, str],
                     progress_cb: Optional[Callable] = None) -> int:
    """replacements: {filename -> replacement_value}"""
    files = get_source_files(folder_path)
    done  = 0
    t0    = time.perf_counter()

    def _do(fp: Path):
        if fp.name not in replacements:
            return
        val = replacements[fp.name]
        ft0 = time.perf_counter()
        (_replace_csv if fp.suffix.lower() == '.csv' else _replace_xlsx)(fp, val)
        log.info(f"[А]  {fp.name}  →  «{val}»  |  {time.perf_counter()-ft0:.2f}с")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futs = {pool.submit(_do, f): f for f in files}
        for fut in as_completed(futs):
            fut.result(); done += 1
            if progress_cb: progress_cb(done, len(files))

    log.info(f"[А]  итого: {time.perf_counter()-t0:.2f}с")
    return done


# ══════════════════════════════════════════════════════════════════════════
#  Сценарий Б: добавить колонку 'dang'
# ══════════════════════════════════════════════════════════════════════════

def _mark_dang_csv(fp: Path, hosts: Set[str]) -> None:
    emp = '' in hosts
    def tr(chunk):
        hcol = find_host_col_df(chunk)
        if hcol:
            chunk['dang'] = chunk[hcol].apply(
                lambda h: 'dang' if (
                    (pd.isna(h) and emp) or
                    (not pd.isna(h) and str(h) in hosts)
                ) else '')
        return chunk
    _rewrite_csv(fp, tr)


def _mark_dang_xlsx(fp: Path, hosts: Set[str]) -> None:
    emp = '' in hosts
    def new_hdr(hdr):
        return hdr if 'dang' in hdr else hdr + ['dang']
    def tr(row, hdr):
        hi   = find_host_idx(hdr)
        hv   = '' if (hi is None or row[hi] is None) else str(row[hi])
        mark = 'dang' if ((not hv and emp) or hv in hosts) else ''
        sh   = hdr
        if 'dang' in sh:
            row[sh.index('dang')] = mark
        else:
            row.append(mark)
        return row
    _rewrite_xlsx(fp, tr, new_hdr)


def scenario_b_apply(folder_path: str,
                     checked: Set[Tuple[str, str]],
                     progress_cb: Optional[Callable] = None) -> int:
    """checked: set of (filename, host_internal)"""
    fname_hosts: dict = defaultdict(set)
    for fn, h in checked: fname_hosts[fn].add(h)
    files = get_source_files(folder_path)
    done  = 0
    t0    = time.perf_counter()

    def _do(fp: Path):
        if fp.name not in fname_hosts: return
        hs  = fname_hosts[fp.name]
        ft0 = time.perf_counter()
        (_mark_dang_csv if fp.suffix.lower() == '.csv'
         else _mark_dang_xlsx)(fp, hs)
        log.info(f"[Б]  {fp.name}  |  {len(hs)} хост(ов)  |  "
                 f"{time.perf_counter()-ft0:.2f}с")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futs = {pool.submit(_do, f): f for f in files}
        for fut in as_completed(futs):
            fut.result(); done += 1
            if progress_cb: progress_cb(done, len(files))

    log.info(f"[Б]  итого: {time.perf_counter()-t0:.2f}с")
    return done


# ══════════════════════════════════════════════════════════════════════════
#  Удаление строк с выбранными хостами
# ══════════════════════════════════════════════════════════════════════════

def _delete_csv(fp: Path, hosts: Set[str]) -> None:
    emp = '' in hosts
    def tr(chunk):
        hcol = find_host_col_df(chunk)
        if hcol:
            mask = ~chunk[hcol].apply(
                lambda h: (pd.isna(h) and emp) or
                          (not pd.isna(h) and str(h) in hosts))
            chunk = chunk[mask]
        return chunk if not chunk.empty else None
    _rewrite_csv(fp, tr)


def _delete_xlsx(fp: Path, hosts: Set[str]) -> None:
    emp = '' in hosts
    def tr(row, hdr):
        hi = find_host_idx(hdr)
        hv = '' if (hi is None or row[hi] is None) else str(row[hi])
        if (not hv and emp) or hv in hosts:
            return None   # удалить строку
        return row
    _rewrite_xlsx(fp, tr)


def delete_hosts_apply(folder_path: str,
                       checked: Set[Tuple[str, str]],
                       progress_cb: Optional[Callable] = None) -> int:
    """checked: set of (filename, host_internal)"""
    fname_hosts: dict = defaultdict(set)
    for fn, h in checked: fname_hosts[fn].add(h)
    files = get_source_files(folder_path)
    done  = 0
    t0    = time.perf_counter()

    def _do(fp: Path):
        if fp.name not in fname_hosts: return
        hs  = fname_hosts[fp.name]
        ft0 = time.perf_counter()
        (_delete_csv if fp.suffix.lower() == '.csv'
         else _delete_xlsx)(fp, hs)
        log.info(f"[Del] {fp.name}  |  {len(hs)} хост(ов)  |  "
                 f"{time.perf_counter()-ft0:.2f}с")

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
        futs = {pool.submit(_do, f): f for f in files}
        for fut in as_completed(futs):
            fut.result(); done += 1
            if progress_cb: progress_cb(done, len(files))

    log.info(f"[Del] итого: {time.perf_counter()-t0:.2f}с")
    return done


# ══════════════════════════════════════════════════════════════════════════
#  GUI — обработчик лога
# ══════════════════════════════════════════════════════════════════════════

class _GUILogHandler(logging.Handler):
    _CLR = {logging.DEBUG: '#888', logging.INFO: '#111',
            logging.WARNING: '#b07000', logging.ERROR: '#c00',
            logging.CRITICAL: '#900'}

    def __init__(self, widget):
        super().__init__()
        self.w = widget
        self.setFormatter(logging.Formatter(
            '%(asctime)s  %(levelname)-8s  %(message)s', datefmt='%H:%M:%S'))

    def emit(self, r):
        msg, lvl, clr = self.format(r) + '\n', r.levelname, self._CLR.get(r.levelno, '#111')
        def _w():
            self.w.configure(state='normal')
            self.w.insert('end', msg, lvl)
            self.w.tag_config(lvl, foreground=clr)
            self.w.see('end')
            self.w.configure(state='disabled')
        self.w.after(0, _w)


# ══════════════════════════════════════════════════════════════════════════
#  GUI — диалог ввода замен (Сценарий А)
# ══════════════════════════════════════════════════════════════════════════

class ReplacementDialog(tk.Toplevel):
    """
    Модальный диалог: для каждого файла — поле ввода замены хостов.
    Можно заполнить не все поля — незаполненные файлы будут пропущены
    с предупреждением.
    """

    def __init__(self, parent: tk.Tk, filenames: List[str]):
        super().__init__(parent)
        self.title('Сценарий А — Замена хостов')
        self.resizable(True, True)
        self.result: Optional[Dict[str, str]] = None
        self._vars: Dict[str, tk.StringVar] = {}

        self._build(filenames)
        self.update_idletasks()
        w = 680
        h = min(110 + len(filenames) * 36 + 70, 580)
        self.geometry(f'{w}x{h}')
        self.minsize(520, 260)
        self.transient(parent)
        self.grab_set()
        self.wait_window()

    def _build(self, filenames: List[str]) -> None:
        # ── Заголовок ──────────────────────────────────────────────────────
        ttk.Label(self,
                  text='Введите название для замены хостов в каждом файле:',
                  font=('Arial', 10), padding=(12, 10, 12, 2)).pack(anchor='w')
        ttk.Label(self,
                  text='Незаполненные файлы будут пропущены (появится предупреждение).',
                  font=('Arial', 9), foreground='#666',
                  padding=(12, 0, 12, 6)).pack(anchor='w')

        # ── Кнопки — пакуем ПЕРВЫМИ с side='bottom', чтобы всегда были видны
        ttk.Separator(self, orient='horizontal').pack(
            side='bottom', fill='x', padx=10, pady=(4, 0))
        bfr = ttk.Frame(self, padding=(10, 6, 10, 10))
        bfr.pack(side='bottom', fill='x')
        ttk.Button(bfr, text='Отмена',
                   command=self.destroy).pack(side='right', padx=(4, 0))
        self._ok = ttk.Button(bfr, text='✔  Применить',
                               command=self._ok_clicked)
        self._ok.pack(side='right')

        # ── Прокручиваемая область ─────────────────────────────────────────
        wrap   = ttk.Frame(self)
        wrap.pack(fill='both', expand=True, padx=10, pady=(0, 4))
        canvas = tk.Canvas(wrap, borderwidth=0, highlightthickness=0)
        scroll = ttk.Scrollbar(wrap, orient='vertical', command=canvas.yview)
        inner  = ttk.Frame(canvas)
        canvas.configure(yscrollcommand=scroll.set)
        canvas.pack(side='left', fill='both', expand=True)
        scroll.pack(side='right', fill='y')
        fid = canvas.create_window((0, 0), window=inner, anchor='nw')
        inner.bind('<Configure>',
                   lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind('<Configure>',
                    lambda e: canvas.itemconfig(fid, width=e.width))

        for fname in filenames:
            row = ttk.Frame(inner, padding=(6, 3, 6, 3))
            row.pack(fill='x')
            ttk.Label(row, text=fname, anchor='w',
                      font=('Consolas', 9), width=40).pack(side='left')
            ttk.Label(row, text='→', font=('Arial', 10)).pack(side='left', padx=4)
            var = tk.StringVar()
            var.trace_add('write', lambda *_: self._validate())
            self._vars[fname] = var
            ttk.Entry(row, textvariable=var, font=('Arial', 10),
                      width=28).pack(side='left', fill='x', expand=True)

        self._validate()

    def _validate(self) -> None:
        filled  = sum(1 for v in self._vars.values() if v.get().strip())
        total   = len(self._vars)
        if filled == 0:
            self._ok.configure(text='✔  Применить')
            self._ok.state(['disabled'])
        elif filled == total:
            self._ok.configure(text='✔  Применить ко всем')
            self._ok.state(['!disabled'])
        else:
            self._ok.configure(text=f'✔  Применить к {filled} из {total}')
            self._ok.state(['!disabled'])

    def _ok_clicked(self) -> None:
        filled  = {f: v.get().strip() for f, v in self._vars.items()
                   if v.get().strip()}
        skipped = [f for f, v in self._vars.items() if not v.get().strip()]

        if skipped:
            show = skipped[:8]
            tail = f'\n  … и ещё {len(skipped)-8}' if len(skipped) > 8 else ''
            msg  = (f'Не заполнено {len(skipped)} файл(ов) — '
                    f'они не будут изменены:\n\n' +
                    '\n'.join(f'  • {f}' for f in show) + tail +
                    '\n\nПродолжить?')
            if not messagebox.askyesno('Внимание', msg, parent=self):
                return

        self.result = filled
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════
#  GUI — главное приложение
# ══════════════════════════════════════════════════════════════════════════

class App:
    _CHK_ON  = '☑'
    _CHK_OFF = '☐'

    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title('Обработка таблиц')
        self.root.geometry('1100x740')
        self.root.minsize(820, 560)

        self._folder: str = ''
        # {filename: [host_internal, ...]} — все данные
        self._data: Dict[str, List[str]] = {}
        # {child_iid: (filename, host_internal)} — видимые дочерние элементы
        self._child_data: Dict[str, Tuple[str, str]] = {}
        # {parent_iid: filename} — видимые родительские элементы
        self._parent_data: Dict[str, str] = {}
        # выбранные пары (filename, host_internal)
        self._checked: Set[Tuple[str, str]] = set()

        self._busy       = False
        self._filter_job: Optional[str] = None

        self._build_ui()

    # ── Построение интерфейса ─────────────────────────────────────────────

    def _build_ui(self) -> None:
        style = ttk.Style()
        try:
            style.theme_use('clam')
        except tk.TclError:
            pass
        style.configure('Action.TButton', font=('Arial', 10, 'bold'), padding=7)
        style.configure('Del.TButton',    font=('Arial', 10, 'bold'), padding=7,
                        foreground='#cc0000')

        self._nb = ttk.Notebook(self.root)
        self._nb.pack(fill='both', expand=True, padx=6, pady=6)

        self._tab1 = ttk.Frame(self._nb)
        self._tab2 = ttk.Frame(self._nb)
        self._nb.add(self._tab1, text='  📂  Обработка файлов  ')
        self._nb.add(self._tab2, text='  🔍  Управление хостами  ')
        self._nb.tab(1, state='disabled')

        self._build_tab1()
        self._build_tab2()

    # ── Вкладка 1 ─────────────────────────────────────────────────────────

    def _build_tab1(self) -> None:
        t = self._tab1

        frm = ttk.LabelFrame(t, text=' Папка с исходными файлами ', padding=10)
        frm.pack(fill='x', padx=12, pady=(12, 6))
        self._fvar = tk.StringVar()
        ttk.Entry(frm, textvariable=self._fvar,
                  font=('Arial', 10)).pack(side='left', fill='x',
                                            expand=True, padx=(0, 6))
        ttk.Button(frm, text='📁  Выбрать папку…',
                   command=self._browse).pack(side='left')

        bfr = ttk.Frame(t, padding=(12, 0))
        bfr.pack(fill='x')
        self._btn_run = ttk.Button(bfr, text='▶  Запустить обработку файлов',
                                    style='Action.TButton',
                                    command=self._run_phase1)
        self._btn_run.pack(side='left', padx=(0, 8))
        self._btn_load = ttk.Button(
            bfr, text='📋  Загрузить существующий hosts_result.xlsx',
            command=self._load_existing)
        self._btn_load.pack(side='left')

        pfr = ttk.Frame(t, padding=(12, 6, 12, 0))
        pfr.pack(fill='x')

        # Строка операции (показывает что сейчас делается)
        self._p1_op = tk.StringVar(value='')
        self._p1_oplbl = ttk.Label(pfr, textvariable=self._p1_op,
                                    font=('Arial', 9, 'italic'),
                                    foreground='#1565c0', anchor='w')
        self._p1_oplbl.pack(fill='x', pady=(0, 2))

        # Вертикальный бар: сначала indeterminate, потом determinate по файлам
        self._p1v = tk.DoubleVar(value=0)
        self._p1bar = ttk.Progressbar(pfr, variable=self._p1v,
                                       mode='determinate', maximum=100)
        self._p1bar.pack(fill='x')
        self._p1lbl = tk.StringVar(value='')
        ttk.Label(pfr, textvariable=self._p1lbl,
                  font=('Arial', 9), anchor='w').pack(fill='x', pady=(2, 0))

        lfr = ttk.LabelFrame(t, text=' Лог обработки ', padding=6)
        lfr.pack(fill='both', expand=True, padx=12, pady=(8, 12))
        self._log = scrolledtext.ScrolledText(
            lfr, height=14, state='disabled',
            font=('Consolas', 9), wrap='word', bg='#fafafa')
        self._log.pack(fill='both', expand=True)
        log.addHandler(_GUILogHandler(self._log))

    # ── Вкладка 2 ─────────────────────────────────────────────────────────

    def _build_tab2(self) -> None:
        t = self._tab2

        # Фильтр
        ffr = ttk.LabelFrame(t, text=' Поиск ', padding=(10, 7))
        ffr.pack(fill='x', padx=12, pady=(10, 4))
        ttk.Label(ffr, text='Поиск по файлу или хосту:',
                  font=('Arial', 10)).pack(side='left')
        self._qv = tk.StringVar()
        self._qv.trace_add('write', lambda *_: self._schedule_filter())
        ttk.Entry(ffr, textvariable=self._qv,
                  width=38, font=('Arial', 10)).pack(side='left', padx=(6, 2))
        ttk.Button(ffr, text='✕', width=3,
                   command=lambda: self._qv.set('')).pack(side='left', padx=(0, 16))
        ttk.Separator(ffr, orient='vertical').pack(side='left', fill='y', padx=4)
        ttk.Button(ffr, text='☑  Выбрать все видимые',
                   command=self._check_all).pack(side='left', padx=4)
        ttk.Button(ffr, text='☐  Снять всё',
                   command=self._uncheck_all).pack(side='left', padx=4)

        # Дерево
        tfr = ttk.Frame(t, padding=(12, 0, 12, 0))
        tfr.pack(fill='both', expand=True)

        self._tree = ttk.Treeview(tfr, columns=('chk',),
                                   show='tree headings', selectmode='browse')
        self._tree.heading('#0',  text='Имя файла / Хост')
        self._tree.heading('chk', text='Выбрано')
        self._tree.column('#0',   width=760, minwidth=300, stretch=True, anchor='w')
        self._tree.column('chk',  width=90,  minwidth=70,  stretch=False, anchor='center')

        vsb = ttk.Scrollbar(tfr, orient='vertical',   command=self._tree.yview)
        hsb = ttk.Scrollbar(tfr, orient='horizontal', command=self._tree.xview)
        self._tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        self._tree.pack(fill='both', expand=True)

        self._tree.bind('<Button-1>', self._on_click)
        self._tree.bind('<space>',    self._on_space)

        # Теги оформления
        self._tree.tag_configure('file',      font=('Arial', 10, 'bold'),
                                  background='#e8eaf6')
        self._tree.tag_configure('file_part', font=('Arial', 10, 'bold'),
                                  background='#e8eaf6', foreground='#1565c0')
        self._tree.tag_configure('file_all',  font=('Arial', 10, 'bold'),
                                  background='#c8e6c9', foreground='#1b5e20')
        self._tree.tag_configure('checked',   background='#c8e6c9',
                                  foreground='#1b5e20')
        self._tree.tag_configure('unchecked', background='',
                                  foreground='')
        self._tree.tag_configure('chk_emp',   background='#c8e6c9',
                                  foreground='#5d7a5d')
        self._tree.tag_configure('unk_emp',   background='',
                                  foreground='#999999')

        # ── Статус (всегда виден) ─────────────────────────────────────────
        ttk.Separator(t, orient='horizontal').pack(fill='x', padx=12, pady=(4, 0))
        self._stv = tk.StringVar(value='—')
        self._stlbl = ttk.Label(t, textvariable=self._stv,
                                 font=('Arial', 9), anchor='w', padding=(12, 3))
        self._stlbl.pack(fill='x')

        # ── Панель прелоадера ─────────────────────────────────────────────
        # Внешний фрейм всегда в иерархии (высота 0 когда скрыт).
        # Внутренний показывается/скрывается через pack/pack_forget.
        self._busy_outer = ttk.Frame(t)
        self._busy_outer.pack(fill='x', padx=12)

        self._busy_inner = ttk.Frame(self._busy_outer, padding=(0, 4, 0, 4))
        # Строка 1: прогрессбар
        self._busy_pv  = tk.DoubleVar(value=0)
        self._busy_bar = ttk.Progressbar(
            self._busy_inner, variable=self._busy_pv,
            mode='indeterminate', maximum=100)
        self._busy_bar.pack(fill='x')
        # Строка 2: название операции
        self._busy_op  = tk.StringVar(value='')
        ttk.Label(self._busy_inner, textvariable=self._busy_op,
                  font=('Arial', 9, 'bold'), foreground='#1565c0',
                  anchor='w').pack(fill='x', pady=(3, 0))
        # Строка 3: детали (текущий файл)
        self._busy_det = tk.StringVar(value='')
        ttk.Label(self._busy_inner, textvariable=self._busy_det,
                  font=('Arial', 8, 'italic'), foreground='#555',
                  anchor='w').pack(fill='x')
        # (внутренний фрейм пока не packed — скрыт)

        # Кнопки действий
        ttk.Separator(t, orient='horizontal').pack(fill='x', padx=12)
        bot = ttk.Frame(t, padding=(12, 8, 12, 10))
        bot.pack(fill='x')
        bot.columnconfigure(0, weight=1)
        bot.columnconfigure(1, weight=1)
        bot.columnconfigure(2, weight=1)

        self._btn_a = ttk.Button(
            bot,
            text='Нет интересных хостов\n→ Задать замену  (Сцен. А)',
            style='Action.TButton', command=self._run_scenario_a)
        self._btn_a.grid(row=0, column=0, sticky='ew', padx=(0, 3), ipady=4)

        self._btn_b = ttk.Button(
            bot,
            text='✔  Отметить выбранные\n    как "dang"  (Сцен. Б)',
            style='Action.TButton', command=self._run_scenario_b)
        self._btn_b.grid(row=0, column=1, sticky='ew', padx=3, ipady=4)

        self._btn_del = ttk.Button(
            bot,
            text='🗑  Удалить строки с\n    выбранными хостами',
            style='Del.TButton', command=self._run_delete)
        self._btn_del.grid(row=0, column=2, sticky='ew', padx=(3, 0), ipady=4)

    # ── Логика Tab 1 ──────────────────────────────────────────────────────

    def _browse(self) -> None:
        path = filedialog.askdirectory(title='Выберите папку с файлами')
        if not path:
            return
        self._fvar.set(path)
        self._folder = path
        out = Path(path) / OUTPUT_FILENAME
        if out.exists():
            log.info(f'Найден существующий файл результатов: {out.name}')
            self._load_hosts_from(out)

    def _run_phase1(self) -> None:
        folder = self._fvar.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showerror('Ошибка', 'Укажите корректный путь к папке.')
            return
        if self._busy:
            return

        # ── Проверка блокировок ДО начала обработки ────────────────────────
        self._p1lbl.set('Проверка доступа к файлам…')
        self.root.update_idletasks()
        locked = check_file_locks(folder)
        if locked:
            self._p1lbl.set('')
            list_txt = '\n'.join(f'  • {f}' for f in locked)
            messagebox.showerror(
                'Файлы заняты',
                f'Следующие файлы открыты в другой программе '
                f'(закройте их и повторите):\n\n{list_txt}')
            return

        self._folder = folder
        self._set_busy1(True)
        self._p1v.set(0)
        self._p1lbl.set('')
        self._p1_op.set('⚙  Удаление лишних столбцов и сбор хостов…')
        # Запускаем бар в indeterminate до первого завершённого файла
        self._p1bar.configure(mode='indeterminate')
        self._p1bar.start(12)

        def run():
            try:
                rows = run_phase1(folder, self._p1cb)
                out  = Path(folder) / OUTPUT_FILENAME
                save_hosts_xlsx(rows, out)
                self.root.after(0, lambda: self._on_p1_done(rows, out))
            except Exception as exc:
                msg = str(exc)
                self.root.after(0, lambda: [
                    self._set_busy1(False),
                    self._p1bar.stop(),
                    self._p1_op.set('✗  Ошибка обработки'),
                    self._p1_oplbl.configure(foreground='#c62828'),
                    messagebox.showerror('Ошибка', msg)])
        threading.Thread(target=run, daemon=True).start()

    def _p1cb(self, done: int, total: int, fname: str = '') -> None:
        pct = done / total * 100 if total else 100
        d, t, fn = done, total, fname
        def _u():
            # Первый завершённый файл — переключаемся на determinate
            self._p1bar.stop()
            self._p1bar.configure(mode='determinate')
            self._p1v.set(pct)
            short = fn[:48] + '…' if len(fn) > 51 else fn
            self._p1lbl.set(f'[{d}/{t}]  {short}  ({pct:.0f}%)')
        self.root.after(0, _u)

    def _on_p1_done(self, rows: list, out: Path) -> None:
        self._set_busy1(False)
        self._p1bar.stop()
        self._p1bar.configure(mode='determinate')
        self._p1v.set(100)
        self._p1_op.set('✓  Обработка завершена')
        self._p1_oplbl.configure(foreground='#2e7d32')
        self._load_hosts_from(out)
        self._nb.tab(1, state='normal')
        self._nb.select(1)

    def _load_existing(self) -> None:
        folder = self._fvar.get().strip()
        if not folder or not Path(folder).is_dir():
            messagebox.showerror('Ошибка', 'Укажите корректный путь к папке.')
            return
        self._folder = folder
        out = Path(folder) / OUTPUT_FILENAME
        if not out.exists():
            messagebox.showerror('Файл не найден',
                                  f'{OUTPUT_FILENAME} не найден в:\n{folder}')
            return
        self._load_hosts_from(out)
        self._nb.tab(1, state='normal')
        self._nb.select(1)

    def _set_busy1(self, busy: bool) -> None:
        self._busy = busy
        for b in (self._btn_run, self._btn_load):
            b.state(('disabled',) if busy else ('!disabled',))

    # ── Загрузка хостов ───────────────────────────────────────────────────

    def _load_hosts_from(self, path: Path) -> None:
        # Мини-индикатор на время чтения Excel (синхронный, но быстрый)
        old_stv = self._stv.get()
        self._stv.set(f'⟳  Загрузка {path.name}…')
        self.root.update_idletasks()
        try:
            df = pd.read_excel(path)
            lc = 'Название файла' if 'Название файла' in df.columns else df.columns[-2]
            hc = 'Имя хоста'     if 'Имя хоста'     in df.columns else df.columns[-1]

            raw: Dict[str, List[str]] = defaultdict(list)
            for fname, host in zip(df[lc].fillna('').astype(str),
                                    df[hc].fillna('').astype(str)):
                raw[fname].append(host)

            # Сортировка: имена файлов A→Z, хосты A→Z, пустые в конце
            self._data = {
                fn: sorted(hs, key=lambda h: (not h, h.lower()))
                for fn, hs in sorted(raw.items(), key=lambda x: x[0].lower())
            }
            self._checked.clear()
            self._child_data.clear()
            self._parent_data.clear()
            self._apply_filter()
            log.info(f'Загружено: {len(self._data)} файлов, '
                     f'{sum(len(v) for v in self._data.values())} хостов')
        except Exception as exc:
            self._stv.set(old_stv)
            log.error(f'Ошибка загрузки: {exc}', exc_info=True)
            messagebox.showerror('Ошибка загрузки', str(exc))

    # ── Фильтрация и построение дерева ────────────────────────────────────

    def _schedule_filter(self) -> None:
        if self._filter_job:
            self.root.after_cancel(self._filter_job)
        self._filter_job = self.root.after(220, self._apply_filter)

    def _apply_filter(self) -> None:
        qry = self._qv.get().strip().lower()

        self._tree.delete(*self._tree.get_children())
        self._child_data.clear()
        self._parent_data.clear()

        for fname, hosts in self._data.items():
            fname_match = qry and qry in fname.lower()

            if qry:
                matching = [h for h in hosts
                            if fname_match or
                               qry in (EMPTY_DISPLAY if not h else h).lower()]
            else:
                matching = hosts

            if not matching:
                continue

            total = len(matching)
            sel   = sum(1 for h in matching if (fname, h) in self._checked)
            cnt   = f'{sel}/{total}'
            tag   = 'file_all' if sel == total else \
                    'file_part' if sel > 0 else 'file'

            p_iid = self._tree.insert(
                '', 'end',
                text=fname, values=(cnt,),
                open=bool(qry),
                tags=(tag,))
            self._parent_data[p_iid] = fname

            for host in matching:
                chk   = (fname, host) in self._checked
                disp  = EMPTY_DISPLAY if not host else host
                is_e  = not host
                ctag  = ('chk_emp' if is_e else 'checked'  ) if chk else \
                        ('unk_emp' if is_e else 'unchecked')
                c_iid = self._tree.insert(
                    p_iid, 'end',
                    text=disp,
                    values=(self._CHK_ON if chk else self._CHK_OFF,),
                    tags=(ctag,))
                self._child_data[c_iid] = (fname, host)

        self._update_status()

    # ── Взаимодействие с деревом ──────────────────────────────────────────

    def _on_click(self, event: tk.Event) -> None:
        if self._tree.identify_region(event.x, event.y) == 'tree':
            return  # клик на треугольник разворачивания — не трогаем
        iid = self._tree.identify_row(event.y)
        if not iid:
            return
        if iid in self._child_data:
            self._toggle_child(iid)
        elif iid in self._parent_data:
            self._toggle_parent(iid)

    def _on_space(self, event: tk.Event) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid in self._child_data:
            self._toggle_child(iid)
        elif iid in self._parent_data:
            self._toggle_parent(iid)

    def _toggle_child(self, c_iid: str) -> None:
        fname, host = self._child_data[c_iid]
        key = (fname, host)
        chk = key not in self._checked
        if chk:
            self._checked.add(key)
        else:
            self._checked.discard(key)
        is_e = not host
        tag  = ('chk_emp' if is_e else 'checked'  ) if chk else \
               ('unk_emp' if is_e else 'unchecked')
        self._tree.item(c_iid,
                         values=(self._CHK_ON if chk else self._CHK_OFF,),
                         tags=(tag,))
        self._update_parent_tag(self._tree.parent(c_iid))
        self._update_status()

    def _toggle_parent(self, p_iid: str) -> None:
        fname    = self._parent_data[p_iid]
        children = [c for c in self._tree.get_children(p_iid)
                    if c in self._child_data]
        keys     = [self._child_data[c] for c in children]
        all_sel  = all(k in self._checked for k in keys)

        for c, key in zip(children, keys):
            _, host = key
            is_e = not host
            if all_sel:
                self._checked.discard(key)
                tag = 'unk_emp' if is_e else 'unchecked'
                val = self._CHK_OFF
            else:
                self._checked.add(key)
                tag = 'chk_emp' if is_e else 'checked'
                val = self._CHK_ON
            self._tree.item(c, values=(val,), tags=(tag,))

        self._update_parent_tag(p_iid)
        self._update_status()

    def _update_parent_tag(self, p_iid: str) -> None:
        if not p_iid:
            return
        children = [c for c in self._tree.get_children(p_iid)
                    if c in self._child_data]
        total = len(children)
        sel   = sum(1 for c in children
                    if self._child_data[c] in self._checked)
        cnt   = f'{sel}/{total}'
        tag   = 'file_all' if sel == total else \
                'file_part' if sel > 0 else 'file'
        self._tree.item(p_iid, values=(cnt,), tags=(tag,))

    def _check_all(self) -> None:
        for c_iid, (fname, host) in self._child_data.items():
            self._checked.add((fname, host))
            is_e = not host
            self._tree.item(c_iid,
                             values=(self._CHK_ON,),
                             tags=('chk_emp' if is_e else 'checked',))
        for p_iid in self._parent_data:
            self._update_parent_tag(p_iid)
        self._update_status()

    def _uncheck_all(self) -> None:
        for c_iid, (fname, host) in self._child_data.items():
            self._checked.discard((fname, host))
            is_e = not host
            self._tree.item(c_iid,
                             values=(self._CHK_OFF,),
                             tags=('unk_emp' if is_e else 'unchecked',))
        for p_iid in self._parent_data:
            self._update_parent_tag(p_iid)
        self._update_status()

    def _update_status(self) -> None:
        all_f = len(self._data)
        all_h = sum(len(v) for v in self._data.values())
        vis_f = len(self._parent_data)
        vis_h = len(self._child_data)
        sel   = len(self._checked)
        flt   = (f'  |  Показано: {vis_f} файлов / {vis_h} хостов'
                 if vis_h != all_h else '')
        self._stv.set(f'Файлов: {all_f}  Хостов: {all_h}{flt}   |   Выбрано: {sel}')

    # ── Состояние «занят» ─────────────────────────────────────────────────

    def _set_busy2(self, busy: bool, operation: str = '') -> None:
        """
        busy=True  → показывает прелоадер с названием операции (indeterminate).
        busy=False → скрывает прелоадер, восстанавливает статус.
        """
        self._busy = busy
        for b in (self._btn_a, self._btn_b, self._btn_del):
            b.state(('disabled',) if busy else ('!disabled',))

        if busy:
            self._busy_pv.set(0)
            self._busy_op.set(f'⚙  {operation}' if operation else '⚙  Обработка…')
            self._busy_det.set('')
            self._busy_bar.configure(mode='indeterminate')
            self._busy_bar.start(12)
            self._busy_inner.pack(fill='x')
        else:
            self._busy_bar.stop()
            self._busy_inner.pack_forget()
            self._update_status()

    def _p2cb(self, done: int, total: int, fname: str = '') -> None:
        """
        Вызывается из фонового потока после завершения каждого файла.
        Переключает бар в determinate-режим и показывает детали.
        """
        pct = done / total * 100 if total else 100
        d, t, fn = done, total, fname
        def _u():
            # Первый файл завершён — переключаем на determinate
            self._busy_bar.stop()
            self._busy_bar.configure(mode='determinate')
            self._busy_pv.set(pct)
            short = fn[:52] + '…' if len(fn) > 55 else fn
            self._busy_det.set(
                f'Завершён [{d}/{t}]: {short}  ({pct:.0f}%)' if fn
                else f'Файлов: {d} / {t}  ({pct:.0f}%)')
        self.root.after(0, _u)

    # ── Проверка папки ────────────────────────────────────────────────────

    def _ensure_folder(self) -> bool:
        if not self._folder or not Path(self._folder).is_dir():
            messagebox.showerror('Ошибка', 'Папка с исходными файлами не задана.')
            return False
        return True

    # ── Сценарий А ────────────────────────────────────────────────────────

    def _run_scenario_a(self) -> None:
        if self._busy or not self._ensure_folder():
            return
        if not self._data:
            messagebox.showwarning('Нет данных', 'Загрузите список хостов.')
            return

        dlg = ReplacementDialog(self.root, list(self._data.keys()))
        if dlg.result is None:
            return  # пользователь отменил

        replacements = dlg.result
        self._set_busy2(True, 'Замена хостов в исходных файлах (Сцен. А)…')

        def run():
            try:
                n = scenario_a_apply(self._folder, replacements, self._p2cb)
                self.root.after(0, lambda: [
                    self._set_busy2(False),
                    messagebox.showinfo('Готово — Сцен. А',
                                         f'Хосты заменены в {n} файле(ах).')])
            except Exception as exc:
                msg = str(exc)
                self.root.after(0, lambda: [self._set_busy2(False),
                                             messagebox.showerror('Ошибка', msg)])
        threading.Thread(target=run, daemon=True).start()

    # ── Сценарий Б ────────────────────────────────────────────────────────

    def _run_scenario_b(self) -> None:
        if self._busy or not self._ensure_folder():
            return
        if not self._checked:
            messagebox.showwarning('Ничего не выбрано', 'Отметьте хотя бы один хост.')
            return
        n = len(self._checked)
        if not messagebox.askyesno(
            'Сценарий Б — Подтверждение',
            f'Добавить метку "dang" для {n} хоста(-ов)?\n\n⚠ Необратимо.'
        ): return
        self._set_busy2(True, f'Добавление метки "dang" для {n} хоста(-ов) (Сцен. Б)…')
        snap = frozenset(self._checked)
        def run():
            try:
                scenario_b_apply(self._folder, snap, self._p2cb)
                self.root.after(0, lambda: [
                    self._set_busy2(False),
                    messagebox.showinfo('Готово — Сцен. Б',
                                         f'"dang" проставлен для {len(snap)} хоста(-ов).')])
            except Exception as exc:
                msg = str(exc)
                self.root.after(0, lambda: [self._set_busy2(False),
                                             messagebox.showerror('Ошибка', msg)])
        threading.Thread(target=run, daemon=True).start()

    # ── Удаление ──────────────────────────────────────────────────────────

    def _run_delete(self) -> None:
        if self._busy or not self._ensure_folder():
            return
        if not self._checked:
            messagebox.showwarning('Ничего не выбрано', 'Отметьте хотя бы один хост.')
            return
        n = len(self._checked)
        if not messagebox.askyesno(
            'Удаление — Подтверждение',
            f'Удалить все строки с {n} выбранным(и) хостом(-ами)?\n\n⚠ Необратимо.'
        ): return
        self._set_busy2(True, f'Удаление строк с {n} выбранным(и) хостом(-ами)…')
        snap = frozenset(self._checked)
        def run():
            try:
                delete_hosts_apply(self._folder, snap, self._p2cb)
                self.root.after(0, lambda: self._after_delete(snap))
            except Exception as exc:
                msg = str(exc)
                self.root.after(0, lambda: [self._set_busy2(False),
                                             messagebox.showerror('Ошибка', msg)])
        threading.Thread(target=run, daemon=True).start()

    def _after_delete(self, deleted: frozenset) -> None:
        # Обновляем _data: убираем удалённые пары
        new_data: Dict[str, List[str]] = {}
        for fname, hosts in self._data.items():
            remaining = [h for h in hosts if (fname, h) not in deleted]
            if remaining:
                new_data[fname] = remaining
        self._data    = new_data
        self._checked -= deleted
        self._apply_filter()
        self._set_busy2(False)
        messagebox.showinfo('Готово — Удаление',
                             f'Удалены строки с хостами: {len(deleted)}.')


# ══════════════════════════════════════════════════════════════════════════
#  Точка входа
# ══════════════════════════════════════════════════════════════════════════

def main() -> None:
    root = tk.Tk()
    App(root)
    root.mainloop()


if __name__ == '__main__':
    main()
    